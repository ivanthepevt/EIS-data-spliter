from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog as fd

#root = Tk()
#root.title("EIS file splitter")
#frm = ttk.Frame(root, padding=20)
#frm.grid()

def getnewdir(olddir):
    splitdir = olddir.split("/")
    oldfilename = splitdir[-1]
    newfilename = ""
    splitdir.pop()
    for text in splitdir:
        newfilename = newfilename + text + "/"
    newfilename = newfilename + "processed_" + oldfilename
    return newfilename

filename = fd.askopenfilename()

print("File chosen: " + filename)
indicate = input("Type your indicator format here (for example: Them 10 ul lan thu #) with # represent the timer: ")

indicator = indicate.split("#")
indicator.append(" ")
#print(indicator)

wb = load_workbook(filename)

wb_new = Workbook()

for ws in wb.worksheets:
    F0 = ws["A21"].value
    timer = 0
    ws_0 = wb_new.create_sheet(ws.title)
    current_row_0 = 1
    for row in ws.values:

        # Sheet 0 duplicates all values
        for col, val in enumerate(row, start=1):
            ws_0.cell(row=current_row_0, column=col).value = val #writing new row to the next empty row of the sheet
        current_row_0 = ws_0.max_row + 1

        if row[0] == F0:
            timer += 1
            ws_0.cell(row=current_row_0 - 1, column=11).value = indicator[0] + str(timer-1) + indicator[1]
            ws_new = wb_new.create_sheet(ws.title + " " + str(timer - 1))
            for col, val in enumerate(label_row, start=1):
                ws_new.cell(row=1, column=col).value = val
            #print(row)
        if timer > 0:
            current_row = ws_new.max_row+1
            for col, val in enumerate(row, start=1):
                ws_new.cell(row=current_row, column=col).value = val #writing new row to the next empty row of the sheet
        else:
            label_row = row


wb_new.save(getnewdir(filename))